from PyPDF2 import PdfFileReader
import os
import openpyxl as op
import time
import smtplib
from tkinter import filedialog
from tkinter import messagebox
from tkinter import *

start = time.time()

class Login:

    #_dateienFolder = os.path.join("Dateien")
    excel = ""
    wrkbk = op.load_workbook(excel)
    sh = wrkbk.active
    pdf = ""
    Documents = os.path.join(pdf)

    personalIds = []
    Vertragsnummern = []
    Names = []


    def __init__(self):
        self.root = Tk()
        self.root.title("Rename PDFs")
        self.root.geometry("1160x920+100+50")
        self.root.resizable(False,False)

        #Hintergund Image
        self.bg = PhotoImage(file="ico\\l.jpg")
        self.bg_image=Label(self.root, image=self.bg).place(x=0,y=-300, relwidth=1, relheight=1)

        #Anmeldungsdaten
        Frame_login = Frame(self.root, bg="white")
        Frame_login.place(x=420, y=300, width=500, height=550)

        #Title
        title = Label(Frame_login, text="DatenVerwaltung", font=("Impact", 35, "bold"), fg="#D4AF37", bg="white").place(x=90,y=30)


        #Path to Excel tabelle
        submit2 = Button(Frame_login, command=self.openExcel, cursor="hand2", text="Excel Path eingeben", bd=0,font=("Goudy old style", 15), bg="#6162FF", fg="white").place(x=90, y=310, width=180, height=40)

        #Path der PDFs
        submit1 = Button(Frame_login, command=self.openPathtoPdfs, cursor="hand2", text="PDFs path eingeben", bd=0,font=("Goudy old style", 15), bg="#6162FF", fg="white").place(x=90, y=380, width=180,height=40)

        #Button command=self.check_function
        submit = Button(Frame_login,command= self.main(), cursor="hand2", text="Start", bd=0, font=("Goudy old style", 15),bg="#6162FF", fg="white").place(x=90, y=450, width=180, height=40)

    def openExcel(self):
        filepath = filedialog.askopenfilename()
        print(filepath)
        Login.excel = filepath
        return filepath

    def openPathtoPdfs(self):
        directory = filedialog.askdirectory()
        print(directory)
        Login.pdf = directory
        return directory



        for i in range(2, Login.sh.max_row + 1):
           for j in range(1, Login.sh.max_column + 1):
            cell_obj = Login.sh.cell(row=i, column=j)
            # print(cell_obj.value , " ",i," ", j ,end=' ')
            # list2 = [cell_obj.value]
            # print(list2)
            if j == 1:
                personalIds.append(cell_obj.value)
            elif j == 2:
                Vertragsnummern.append(cell_obj.value)
            elif j == 3:
                Names.append(cell_obj.value)
            elif j == 4:
                Names[i - 2] += " " + cell_obj.value



    def main(self):
        for root, directory, files in os.walk(Login.Documents):
            for my_file in files:
                old_name = Login.Documents + "\\" + my_file
                temp = open(old_name, 'rb')
                _pdfRead = PdfFileReader(temp)
                page = _pdfRead.getPage(0)
                text = page.extractText()
                temp.close()
                _requestedInfo = self.getInfos(text)
                _organisedData = self.setDataForList(_requestedInfo)
                info_list = self.getInfoList(_organisedData)
                client_id = self.setClientId(info_list[0])
                real_name = self.reverseName(info_list[1])
                _index_in_Names = Login.Names.index(real_name)
                personalnummer = str(Login.personalIds[_index_in_Names])
                vertragsnummer = str(Login.Vertragsnummern[_index_in_Names])
                new_name = Login.Documents + "\\" + personalnummer + "_" + vertragsnummer + "_" + real_name + ".pdf"
                if os.path.isfile(new_name):
                    print("File with the same name already exists !!!")
                else:
                    os.rename(old_name, new_name)
                    print("Successfully changed the name of the file with the name : " + my_file)

        end = time.time()
        print(end - start, " ms")

    def getInfos(self,text):
        wantedInfos = ""
        for i in range(len(text)):
            if text[i] == 'N':
                if self.checkIfKey(text[i] + text[i + 1] + text[i + 2] + text[i + 3]):
                    _curIndex = i + 4
                    while text[_curIndex] != '\n':
                        wantedInfos += text[_curIndex]
                        _curIndex += 1
                    return wantedInfos

    def checkIfKey(self,info):
        return info == "Nr. "

    def setDataForList(self,info):
        Info = ""
        indexOfTrimValue = info.find('-')
        for i in range(len(info)):
            if i == indexOfTrimValue:
                Info += '_'
            elif i < indexOfTrimValue - 1:
                Info += info[i]
            elif i > indexOfTrimValue + 1:
                Info += info[i]
        return Info

    def getInfoList(self,organisedInfo):
        info_list = list()
        curInfo = ""
        for i in range(len(organisedInfo)):
            if organisedInfo[i] != '_':
                curInfo += organisedInfo[i]
                if i == len(organisedInfo) - 1:
                    info_list.append(curInfo)
            else:
                info_list.append(curInfo)
                curInfo = ""
        return info_list

    def setClientId(self,nonProcessedId):
        id = ""
        for i in nonProcessedId:
            if i != '/':
                id += i
        return id

    def reverseName(self,name):
        firstname = ""
        familyName = ""
        _spaceIndex = name.find(' ')
        for i in range(len(name)):
            if i < _spaceIndex:
                familyName += name[i]
            elif i > _spaceIndex:
                firstname += name[i]
        return firstname + " " + familyName

        # if __name__ == '__main__':
        # main()



    def start(self):
        self.root.mainloop